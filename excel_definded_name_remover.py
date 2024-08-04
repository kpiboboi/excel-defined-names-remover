from openpyxl import load_workbook
from openpyxl.workbook.properties import CalcProperties

def remove_named_ranges(input_file_path):
    try:
        # Excel faylini ochish
        workbook = load_workbook(filename=input_file_path)

        # Formulalarni avtomatik qayta hisoblashni o'chirish
        workbook.properties.calcPr = CalcProperties(fullCalcOnLoad=False)

        # O'chiriladigan nomli elementlarning nusxasini yaratish
        named_ranges_to_remove = list(workbook.defined_names)

        # O'chirib bo'lmaydigan nomli elementlar
        failed_to_remove = []

        # Har bir varaqdan barcha nomlangan elementlarni olib tashlash
        for sheet in workbook.sheetnames:
            current_sheet = workbook[sheet]
            for name in named_ranges_to_remove:
                if name in current_sheet.defined_names:
                    try:
                        del current_sheet.defined_names[name]
                    except Exception as e:
                        print(f"Oʻchirishda xatolik yuz berdi '{name}': {e}")
                        failed_to_remove.append(name)

        # Formulalarni avtomatik qayta hisoblash rejimini yoqamiz
        workbook.properties.calcPr = CalcProperties(fullCalcOnLoad=True)

        # O'zgarishlarni o'sha faylni o'zida saqlash
        workbook.save(filename=input_file_path)
        print("✅ Deyalri barcha nomlangan elementlar muvaffaqiyatli o'chirildi! ✅")
        
        # Oʻchirib boʻlmaydigan nomli elementlar roʻyxati
        if failed_to_remove:
            print("O'chirib bo'lmaydigan nomli elementlar:")
            print(failed_to_remove)
    except Exception as e:
        print(f"Xatolik yuz berdi: {e}")

input_file_path = r"C://path/s/s/name.xlsx"
remove_named_ranges(input_file_path)
print("💬 E'tibor bering, bu kod faqat juda ham ko'p nomlangan elementlar bor elementalrni o'chirishga qo'llashingiz mumkin.")
print("💬 Nomli elementlar juda ko'p bo'lsa, Excelni o'zidan o'chira olmasiz. Hattoki, tugmasi ham ishlamaydi.")
print("💬 Bu dastur esa bunday holatlardan halos bo'lishga yordam bera olishi mumkin.")
print("💬 Dastur barcha elementlarni o'chirmaydi. Qolganlarini Excelning o'zidan o'chira olasiz.")